#!/usr/bin/env python3
"""
令和7年度 沖縄県移住定住促進にかかる取り組み状況調査(Excel) を
アプリ用の正規化JSONに変換する。

使い方:
    python scripts/parse_survey.py <survey.xlsx> [--outdir data/generated]

出力:
    municipalities.json  市町村マスタ(コード・名称)
    policy.json          シートA 移住定住施策全般
    housing.json         シートB 定住促進住宅整備(物件単位)
    akiya.json           シートC 空き家利活用に向けた取組
    quality_report.json  名寄せ失敗・数値抽出の要確認箇所
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SHEET_A = "A 移住定住施策全般(R7)"
SHEET_B = "B 定住促進住宅整備(R7)"
SHEET_C = "C 空き家利活用に向けた取組(R7) "   # 末尾スペースは原本どおり

FIRST_DATA_ROW = 5          # 3行目が見出し、4行目は「選択肢保存用」のダミー行


# ---------------------------------------------------------------- utilities
# 個人の連絡先は公開しない。原本 data/raw/survey/latest.xlsx には担当者の携帯番号と
# メールアドレスが入っており、そのまま data/generated/ に出ると GitHub Pages で
# 公開されてしまう。初回コミット(6a5294b)では policy.json を手で伏せ字にしていたが、
# parse_survey.py を再実行すると元に戻ってしまうため変換側に移した。
# **この処理を外さないこと。**
# 括弧でくくられた連絡先だけを対象にする（本部町の「企画商工観光課 / 野崎 真 /
# 0980-47-2702」のような窓口の代表番号は公開してよいので触らない）。
REDACTIONS = [
    (re.compile(r"（\s*[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\s*[)）]"),
     "（メールアドレスは非公開）"),
    (re.compile(r"（\s*0\d{1,4}-\d{1,4}-\d{4}\s*[)）]"), "（非公開）"),
]
# 伏せ字にできなかった個人情報を拾う網。quality_report に出して人が気づけるようにする。
LEAK_PATTERNS = [
    ("email", re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")),
    ("mobile", re.compile(r"0[789]0-\d{4}-\d{4}")),
]


def redact(s):
    """個人の連絡先を伏せる。"""
    for pat, repl in REDACTIONS:
        s = pat.sub(repl, s)
    return s


def norm_text(v):
    """セル値を1行テキストに正規化。空欄はNone。個人の連絡先はここで伏せる。"""
    if v is None:
        return None
    s = str(v).replace("\u3000", " ")
    s = re.sub(r"[\r\n]+", " / ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return redact(s) or None


def norm_choice(v):
    """選択式の値を正規化。'E. 空き家調査は…' のような記号接頭辞を落とす。"""
    s = norm_text(v)
    if s is None:
        return None
    return re.sub(r"^[A-Za-zＡ-Ｚａ-ｚ0-9０-９]{1,2}[.．、)）]\s*", "", s)


def to_bool(v):
    if isinstance(v, bool):
        return v
    s = norm_text(v)
    return s.lower() == "true" if s else False


def fill_merged(ws):
    """結合セルの値を範囲内の全セルに複製する(縦結合の穴埋め)。"""
    for rng in list(ws.merged_cells.ranges):
        anchor = ws.cell(rng.min_row, rng.min_col).value
        ws.unmerge_cells(str(rng))
        if anchor is None:
            continue
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                ws.cell(r, c).value = anchor


def rows_of(ws, name_col):
    """データ行を (行番号, 市町村名, 値リスト) で返す。名称が空の行は前行の続きとみなす。"""
    out, last_name = [], None
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        nm = norm_text(vals[name_col - 1])
        if nm == "選択肢保存用":
            continue
        if nm is None:
            nm = last_name          # 継続行
        else:
            last_name = nm
        if nm is None:
            continue
        out.append((r, nm, vals))
    return out


# ------------------------------------------------------- numeric extraction
# 備考欄の自然文から件数を拾う。確実に取れないものは None にして要確認へ回す。
VACANT_PATTERNS = [
    r"推定空家数?[^0-9]{0,8}([0-9,]+)\s*[件戸]",
    r"推定空き?家(?:件数)?[^0-9]{0,8}([0-9,]+)\s*[件戸]",
    r"空き?家件数[^0-9]{0,4}([0-9,]+)\s*[件戸]",
    r"推定\s*([0-9,]+)\s*[件戸]ほど",
]
BANK_TOTAL_PATTERNS = [
    r"累計登録(?:物件)?数?[^0-9]{0,12}([0-9,]+)\s*件",
    r"登録(?:物件)?件?数[：:]\s*([0-9,]+)\s*件",
    r"累計掲載件数\s*([0-9,]+)\s*件",
]
DEAL_PATTERNS = [
    r"成約済み件数[：:]?\s*([0-9,]+)\s*件",
    r"成約済み[：:]?\s*([0-9,]+)\s*件",
]


ZEN2HAN = str.maketrans("０１２３４５６７８９，．", "0123456789,.")


def ascii_digits(text):
    """全角数字・全角カンマを半角化(抽出専用。表示用テキストは原文のまま残す)。"""
    return text.translate(ZEN2HAN) if text else text


def excel_serial_to_date(text):
    """日付書式が落ちてシリアル値になった欄(例: 46082)をISO日付に戻す。"""
    if text and re.fullmatch(r"\d{5}", text) and 30000 <= int(text) <= 60000:
        from datetime import date, timedelta
        return (date(1899, 12, 30) + timedelta(days=int(text))).isoformat()
    return text


def extract_int(text, patterns):
    text = ascii_digits(text)
    if not text:
        return None
    for p in patterns:
        m = re.search(p, text)
        if m:
            return int(m.group(1).replace(",", ""))
    return None


# 推定空家数の「基準年」。備考欄には無関係の年（助成制度の開始予定など）が
# 混ざるので、拾った数値のすぐ近くだけを見る。範囲内に年が2つ以上あるときは
# どれが基準年か決められないので None（画面では「基準年不明」）に倒す。
# 正規表現を緩めて誤検出を作らないこと（CLAUDE.md 原則4）。
ERA_RE = re.compile(r"(令和|平成|[RH])\s*([0-9]{1,2})\s*年?度?")
ERA_BASE = {"令和": 2018, "R": 2018, "平成": 1988, "H": 1988}
BEFORE_WINDOW, AFTER_WINDOW = 40, 15


def extract_base_year(text, patterns):
    """数値の近傍にある年号を1つだけ拾う。決められなければ None。

    戻り値は (表示用ラベル, 西暦) の組。例: ("令和6年度", 2024)
    """
    text = ascii_digits(text)
    if not text:
        return None, None
    for p in patterns:
        m = re.search(p, text)
        if not m:
            continue
        window = text[max(0, m.start() - BEFORE_WINDOW): m.end() + AFTER_WINDOW]
        found = ERA_RE.findall(window)
        years = {(ERA_BASE[era[0]] if era in ("R", "H") else ERA_BASE[era]) + int(num)
                 for era, num in found}
        if len(years) != 1:      # 0件＝書かれていない、2件以上＝どれか決められない
            return None, None
        era, num = found[0]
        wa = "令和" if era in ("令和", "R") else "平成"
        return f"{wa}{int(num)}年度", years.pop()
    return None, None


def extract_count_field(text):
    """⑬空き住戸数のような「数値のはずだが文章」の列を数値+原文に分ける。"""
    if text is None:
        return None, None
    if re.fullmatch(r"該当なし|なし", text.strip()):
        return 0, text
    m = re.search(r"([0-9,]+)\s*[戸件]?", ascii_digits(text))
    n = int(m.group(1).replace(",", "")) if m else None
    return n, text


# ------------------------------------------------------------------- parsers
def parse_a(ws):
    fields = {
        4: "policy_status", 5: "policy_status_other", 6: "policy_detail", 7: "policy_none_reason",
        8: "desk_status", 9: "desk_status_other", 10: "desk_contact", 11: "desk_none_reason",
        12: "intermediary_status", 13: "intermediary_other", 14: "intermediary_detail",
        15: "intermediary_none_reason",
        16: "jobinfo_status", 17: "jobinfo_other", 18: "jobinfo_detail", 19: "jobinfo_none_reason",
        20: "lot_sale_status", 21: "lot_sale_units", 22: "new_housing_plan", 23: "new_housing_funding",
    }
    choice_cols = {4, 8, 12, 16, 20, 22}
    recs = []
    for r, name, vals in rows_of(ws, 3):
        rec = {
            "code": norm_text(vals[1]),
            "name": name,
            "confirmed": to_bool(vals[0]),
            "note": norm_text(vals[23]),
            "_row": r,
        }
        for col, key in fields.items():
            rec[key] = (norm_choice if col in choice_cols else norm_text)(vals[col - 1])
        recs.append(rec)
    return recs


def parse_b(ws):
    fields = {
        4: "name", 5: "property_type", 6: "units", 7: "floor_area_sqm", 8: "completion_fy",
        9: "occupancy_conditions", 10: "max_tenancy", 11: "funding",
        12: "total_cost_thousand_jpy", 13: "private_finance", 14: "site_origin", 15: "operator",
    }
    choice_cols = {5, 13, 14}
    recs = []
    for r, muni, vals in rows_of(ws, 3):
        rec = {
            "code": norm_text(vals[0]),
            "municipality": muni,
            "confirmed": to_bool(vals[1]),
            "note": norm_text(vals[15]),
            "_row": r,
        }
        for col, key in fields.items():
            rec[key] = (norm_choice if col in choice_cols else norm_text)(vals[col - 1])
        for k in ("floor_area_sqm", "total_cost_thousand_jpy"):
            v = rec[k]
            rec[k + "_num"] = float(v.replace(",", "")) if v and re.fullmatch(r"[0-9,.]+", v) else None
        recs.append(rec)
    return recs


def parse_c(ws):
    """1市町村が複数行(調査種別ごと)に分かれるので、市町村単位に畳む。"""
    fields = {
        5: "consult_desk", 6: "consult_desk_org",
        7: "akiya_plan", 8: "akiya_action_plan", 9: "plan_timing",
        10: "akiya_bank", 11: "statutory_council", 12: "support_corporation",
        13: "subsidy_status", 14: "subsidy_types", 15: "public_housing_vacant",
    }
    choice_cols = {5, 7, 8, 10, 11, 12, 13}
    by_muni = {}
    for r, muni, vals in rows_of(ws, 2):
        rec = by_muni.get(muni)
        if rec is None:
            rec = {
                "municipality": muni,
                "confirmed": to_bool(vals[0]),
                "notes": [],
                "surveys": [],
                "_rows": [],
            }
            for col, key in fields.items():
                rec[key] = (norm_choice if col in choice_cols else norm_text)(vals[col - 1])
            rec["plan_timing"] = excel_serial_to_date(rec["plan_timing"])
            by_muni[muni] = rec
        rec["_rows"].append(r)
        n = norm_text(vals[15])
        if n and n not in rec["notes"]:
            rec["notes"].append(n)
        survey = {"type": norm_choice(vals[2]), "detail": norm_text(vals[3])}
        if (survey["type"] or survey["detail"]) and survey not in rec["surveys"]:
            rec["surveys"].append(survey)

    for rec in by_muni.values():
        rec["note"] = " / ".join(rec.pop("notes")) or None
        note = rec["note"]
        rec["estimated_vacant"] = extract_int(note, VACANT_PATTERNS)
        label, west = extract_base_year(note, VACANT_PATTERNS)
        rec["estimated_vacant_base_year"] = label          # 例 "令和6年度"。取れなければ None
        rec["estimated_vacant_base_year_west"] = west      # 例 2024
        rec["bank_registered_total"] = extract_int(note, BANK_TOTAL_PATTERNS)
        rec["bank_deals_closed"] = extract_int(note, DEAL_PATTERNS)
        n, raw = extract_count_field(rec.get("public_housing_vacant"))
        rec["public_housing_vacant_num"] = n
        rec["public_housing_vacant_raw"] = raw
    return list(by_muni.values())


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--outdir", default="data/generated")
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    for s in (SHEET_A, SHEET_B, SHEET_C):
        if s not in wb.sheetnames:
            sys.exit(f"シートが見つかりません: {s!r}\n実際のシート名: {wb.sheetnames}")
        fill_merged(wb[s])

    policy = parse_a(wb[SHEET_A])
    housing = parse_b(wb[SHEET_B])
    akiya = parse_c(wb[SHEET_C])

    # 市町村マスタ: コードを持つのはシートAだけなので、これを正とする
    name2code = {p["name"]: p["code"] for p in policy if p["code"]}
    municipalities = [{"code": p["code"], "name": p["name"]} for p in policy if p["code"]]

    unmatched = []
    for rec in housing:
        rec["code"] = rec.get("code") or name2code.get(rec["municipality"])
        if not rec["code"]:
            unmatched.append({"sheet": "B", "name": rec["municipality"], "row": rec["_row"]})
    for rec in akiya:
        rec["code"] = name2code.get(rec["municipality"])
        if not rec["code"]:
            unmatched.append({"sheet": "C", "name": rec["municipality"], "rows": rec["_rows"]})

    # 伏せ字の網から漏れた個人情報が無いか、出力全体を最後にもう一度さらう。
    # 新年度の原本で書式が変われば REDACTIONS が効かないことがあるため、
    # 気づかないまま公開しないよう必ず落とす。
    leaks = []
    for label, dataset in (("policy", policy), ("housing", housing), ("akiya", akiya)):
        for rec in dataset:
            for key, val in rec.items():
                if not isinstance(val, str):
                    continue
                for kind, pat in LEAK_PATTERNS:
                    if pat.search(val):
                        leaks.append({"dataset": label, "municipality": rec.get("municipality"),
                                      "field": key, "kind": kind})
    if leaks:
        for x in leaks:
            print(f"  {x['dataset']}.{x['field']} ({x['municipality']}) に{x['kind']}が残っています",
                  file=sys.stderr)
        sys.exit(f"個人情報が伏せられていません（{len(leaks)}件）。"
                 f"REDACTIONS を原本の書式に合わせてから再実行してください。")

    needs_review = []
    for rec in akiya:
        note = rec.get("note")
        if note and rec["estimated_vacant"] is None \
           and re.search(r"空[きー]?家|空家", note) and re.search(r"[0-9]+\s*[件戸]", note):
            needs_review.append({"municipality": rec["municipality"],
                                 "field": "estimated_vacant", "text": note[:160]})
        if rec.get("public_housing_vacant_raw") and rec["public_housing_vacant_num"] is None:
            needs_review.append({"municipality": rec["municipality"],
                                 "field": "public_housing_vacant",
                                 "text": rec["public_housing_vacant_raw"][:160]})

    quality = {
        "counts": {
            "municipalities": len(municipalities),
            "policy_rows": len(policy),
            "housing_properties": len(housing),
            "akiya_municipalities": len(akiya),
        },
        "unconfirmed_policy": [p["name"] for p in policy if not p["confirmed"]],
        "unconfirmed_akiya": [a["municipality"] for a in akiya if not a["confirmed"]],
        "code_join_failures": unmatched,
        "needs_manual_review": needs_review,
    }

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    for fn, obj in {
        "municipalities.json": municipalities,
        "policy.json": policy,
        "housing.json": housing,
        "akiya.json": akiya,
        "quality_report.json": quality,
    }.items():
        (outdir / fn).write_text(json.dumps(obj, ensure_ascii=False, indent=1), encoding="utf-8")
        print(f"wrote {outdir / fn}")

    print("\n--- counts ---")
    print(json.dumps(quality["counts"], ensure_ascii=False))
    print("--- 未確認(シートA) ---", quality["unconfirmed_policy"])
    print("--- 未確認(シートC) ---", quality["unconfirmed_akiya"])
    print("--- コード紐付け失敗 ---", quality["code_join_failures"])
    print(f"--- 要目視 {len(quality['needs_manual_review'])}件 ---")
    for x in quality["needs_manual_review"]:
        print("   ", x["municipality"], "|", x["field"], "|", x["text"][:90])


if __name__ == "__main__":
    main()
